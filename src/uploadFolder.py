import os
import re
from openpyxl import Workbook


# Author: Ryan Freas
# Date: 4/1/2024
# Goal: Search for instances of SQL calls (or any string you are looking for)
# in a directory of your choice, and then export the result to an excel sheet/csv.
#
# Update the os path to the name of the folder you want to search through, then
# update search_pattern with the strings you want to search for.

search_pattern = [r'SqlDataAdapter\("', r'SqlCommand\("', r'"UPDATE ', r'"update ', r'"SELECT ',
                   r'"select ', r'"DELETE ', r'"delete ', r'"INSERT ', r'"insert ', r"<asp:SqlDataSource"]

'''
entries is an array that contains the absolute path of the folder you want to search for, as well as 
a tuple that has the title of your excel sheet, an array of the strings you want to search and match,
and the name of the excel sheet you create.
'''
entries = [
  (
    os.path.abspath("../String-Finder-1.0.0/IREMRedesign"),
    ["Page Template Library: .ascx and c#", 
    search_pattern,
    "IREMRedesign.xlsx"],
  )
  ]
    
def save_to_wb(directory: str, excel_props: tuple) -> None:
    sql_src = handle_all_files(directory, excel_props[1])
    # Don't create excel sheet if no patterns match 
    if len(sql_src) == 0:
        # print 
        dir = (os.listdir(directory))
        return
        # If needed we can swap return for an Exception but for now keep as unreachable
        raise Exception("Your excel sheet will be null. Terminating...")
    wb = Workbook()
    ws = wb.active
    ws.append(["Title: ", excel_props[0]])
    ws.append(["Patterns: ", str(excel_props[1])])
    ws.append(["Files included/excluded: ", '(file.endswith(".cs") or file.endswith(".ascx")) and not (file.endswith(".designer.cs") or file.endswith(".Designer.cs")),  if re.match(r"^/", line.strip()): ', ".cs, .ascx, NOT .designer.cs, and excluding any commented out lines (//)"])
    ws.append(["SQL Calls: ", len(sql_src)])
    currentDir, currentFile = "", ""
    fileCount = 0
    for root, file, line_number, command, pattern in sql_src:
        # Show new folder/path in excel sheet on entry
        if (root != currentDir):
            currentDir = root
            ws.append([""]), ws.append([""])
            ws.append(["PATH:", root[78:]])
            ws.append(["File", "Line Number","Pattern Matched", "Select Command"])
        if (file != currentFile ):
            currentFile = file
            fileCount +=1
        ws.append([file, line_number, pattern, command])

    ws["A5"] = "Unique Files: "
    ws["B5"] = fileCount
    wb.save(excel_props[2])
    print(f"Excel file '{excel_props[2]}' has been created.")
    return

# If we want to exclude SqlCommand("") that aren't 
# a stored procedure we can update the pattern
def handle_all_files(directory, patterns):
    # Searches all folders for files in your directory
    sql_to_excel = []
    for root, folder, files in os.walk(directory):
        for file in files:
            print(file)
            if  (file.endswith(".cs") or file.endswith(".ascx")) and not (file.endswith(".designer.cs") or file.endswith(".Designer.cs")):
                handle_os_walk(patterns, root, file, sql_to_excel)      
    return sql_to_excel

def handle_os_walk(patterns, root, file, sql_to_excel):
    # Searches your files for the specified SQL call pattern
    filepath = os.path.join(root, file)
    # you will get charmap codec can't decode byte XXXX errors if encoding not set properly
    with open(filepath, encoding="utf8") as f:
        try:
            content = f.readlines()
        except:
            print("Invalid byte error in ", file)

        try:
            for line_number, line in enumerate(content, 1):
                for pattern in patterns:
                    # Trying to exclude commented out lines (lines that start with /)
                    if re.match(r'^/', line.strip()): 
                        continue
                    if re.search(pattern, line):
                        sql_to_excel.append(
                                    (root, file, line_number, line.strip(), pattern)
                        )
                        # Break to prevent duplicate entries if diff patterns observed on same line
                        break
        except:
            print("Cannot access local variable content 'UNDEFINED'")
    return 

# Start script and spread entry properties in save_to_wb function
for entry in entries:
    save_to_wb(*entry)